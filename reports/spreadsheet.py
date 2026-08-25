"""The findings as a workbook, with `1/8` still reading `1/8`.

The client's own workflow is spreadsheet-shaped — the checklist that started this project is an
`.xlsx` — so this is the format in which the output gets used rather than re-keyed. That is the
entire reason it exists, and it sets the one hard constraint.

**Every measurement is a text cell.** A spreadsheet is the most eager type-coercer in the toolchain:
give Excel `1/8` in a numeric cell and it stores `0.125`, then shows whatever the column's format
says. The distinction between an eighth of an inch and a decimal approximation of one is what the
whole units layer exists to preserve (ADR-0001), and losing it in the final step would make the
export the least trustworthy artifact the system produces — while looking the most familiar.

So values are written as strings, each cell's number format is pinned to text, and
`tests/test_spreadsheet.py` reads the file back and asserts the stored type. Writing a string is not
enough on its own: openpyxl will happily give a string cell a numeric format, and a consumer that
re-saves the file can then have it converted for them.

**Abstentions are rows.** A `NOT_FOUND` check has no numbers to put in the value columns, and the
tempting thing is to leave the row out. That reproduces exactly the failure `NO_APPLICABLE_RULE` was
invented to stop: a reader scanning for problems sees nothing and reads it as nothing wrong. Every
finding gets a row, and the value columns say what is missing and why.

An abstention is not always empty, and the value columns follow **the calculation, not the outcome**.
A `REVIEW_REQUIRED` raised partway through arithmetic carries a trace, and its comparison and
operands are written out: an abstention raised because two readings disagreed is one whose readings
are the most useful thing in the file.

**Two sheets, because a finding and an operand are different rows.** One row per finding answers
"what did this check conclude?"; one row per traced operand answers "which numbers did it use, and
where did each come from?". Flattening both into one sheet means either repeating the finding on
every operand or dropping operands after the first, and the second is the kind of quiet truncation
that makes a report wrong without looking wrong.

**No expected/observed columns.** They would have to be invented. A `CalculationTrace` records its
operands by name and its `comparison` as text; nothing in it labels one side expected and the other
observed, and guessing from operand order would mislabel every rule whose operands are declared the
other way round. The comparison string is written verbatim instead, and the operand sheet carries
each named value — so a reader can see what was compared without this module asserting which was
which. There is likewise no `item` column: nothing in a finding identifies an item yet, and item
identifiers are B7.3 (#166).

Source: `AGENTS.md` §2.8; ADR-0001 · Design: `docs/DESIGN_PRODUCT.md` §3.3 ·
Verification: ``tests/reports/test_spreadsheet.py``
"""

from __future__ import annotations

import json
from collections.abc import Sequence
from fractions import Fraction
from io import BytesIO
from typing import Final

# openpyxl ships no type stubs, the same situation as reportlab in `reports/redline.py`. Ignored
# per-import rather than repo-wide, so a genuinely untyped import somewhere else still surfaces.
from openpyxl import Workbook  # type: ignore[import-untyped]
from openpyxl.cell.cell import Cell  # type: ignore[import-untyped]
from openpyxl.styles import Font  # type: ignore[import-untyped]
from openpyxl.utils import get_column_letter  # type: ignore[import-untyped]

from units.measurement import Measurement
from verdict.finding import Finding
from verdict.outcomes import is_abstention

__all__ = [
    "FINDINGS_SHEET",
    "FINDING_COLUMNS",
    "OPERANDS_SHEET",
    "OPERAND_COLUMNS",
    "TEXT_FORMAT",
    "UNREADABLE_REFERENCE",
    "exact_text",
    "write_value",
    "write_workbook",
]

#: openpyxl's number format for "leave this alone". Pinned on every value cell.
TEXT_FORMAT: Final = "@"

FINDINGS_SHEET: Final = "Findings"
OPERANDS_SHEET: Final = "Operands"

#: Frozen. Downstream consumers index by position, so inserting a column in the middle silently
#: shifts every one after it — a reader would get tolerances under the severity heading and have no
#: reason to doubt them. Add new columns at the end.
FINDING_COLUMNS: Final = (
    "check",
    "outcome",
    "severity",
    "comparison",
    "difference",
    "tolerance",
    "arithmetic_unit",
    "variant",
    "rule_snapshot",
    "engine_version",
    "evidence_pages",
    "reason",
    "notes",
)

#: One row per operand the calculation used. Frozen for the same reason.
OPERAND_COLUMNS: Final = (
    "check",
    "operand",
    "value",
    "source",
    "evidence_page",
    "evidence_uri",
)

#: What a value column says when a check abstained before producing that value.
#:
#: A word rather than a blank. An empty cell is ambiguous between "nothing to report", "the export
#: dropped it" and "the check was never run", and those want different responses from a reader.
NOT_APPLICABLE: Final = "n/a — check abstained"

#: What a value column says when a decision was reached but the rule declares no such value.
NONE_DECLARED: Final = "none declared"

#: What a value column says when a reference could not be read as a complete citation.
UNREADABLE_REFERENCE: Final = "unreadable reference"


def exact_text(value: object) -> str:
    """Render a value as text that has lost nothing.

    `Fraction` renders as `1/8`, never `0.125`. `Measurement` renders as its exact value with the
    unit it was authored in, because a number without its unit is the other half of the same
    mistake — and it keeps the source token when there is one, so a reader can see what the drawing
    actually said as well as what it was read as.
    """
    if isinstance(value, Measurement):
        rendered = f"{value.exact} {value.unit.value}"
        return f"{rendered} (as written: {value.raw_text})" if value.raw_text else rendered
    if isinstance(value, Fraction):
        return str(value)
    if isinstance(value, tuple):
        return "; ".join(exact_text(item) for item in value)
    if value is None:
        return ""
    return str(value)


def write_value(cell: Cell, value: object) -> None:
    """Write one value as text, and pin the cell's format so it stays text.

    Both halves matter. A spreadsheet that turns `1/8` into `0.125` has discarded the distinction
    the entire units layer exists to preserve — and a string cell left on the `General` format is
    one re-save away from a consumer's tooling doing the conversion on their behalf.
    """
    cell.value = exact_text(value)
    cell.number_format = TEXT_FORMAT


def _decode_reference(reference: str) -> tuple[str, str] | None:
    """(page, document version) for one evidence reference, or `None` if it is not a citation.

    **Both fields or neither.** A reference carrying `page` and no `document_version_id` decodes
    happily and yields "page 1" — of which drawing, nobody can say. Reporting that as provenance
    would put an unfollowable citation in the column a reader uses to go and check, which is worse
    than admitting the reference could not be read: they would look, fail to find it, and doubt the
    finding rather than the export.

    One decoder for both callers, deliberately. Two copies of "what counts as a readable reference"
    is how the operand sheet and the findings sheet come to disagree about the same reference — and
    a reader comparing them would have no way to tell which was right.

    `evidence/gate.py` writes `document_version_id`, `page`, `polygon` and `space`. Only the two
    fields this module reports are required here; a reference with no polygon is still a page in a
    document, and refusing it would drop a usable citation to enforce a field nothing here reads.
    """
    try:
        decoded = json.loads(reference)
        page = decoded["page"]
        document = decoded["document_version_id"]
    except (ValueError, TypeError, KeyError):
        return None
    if isinstance(page, bool) or not isinstance(page, int) or not isinstance(document, str):
        return None
    return (str(page + 1), document)


def _evidence_pages(finding: Finding) -> str:
    """The pages this finding's evidence sits on, for someone holding the drawing set.

    An unreadable reference is reported as such rather than skipped: a reference the export could
    not decode is a fact about the export, and dropping it would leave the row looking like a
    finding with no evidence.
    """
    decoded = [_decode_reference(reference) for reference in finding.evidence_refs]
    return ", ".join(UNREADABLE_REFERENCE if parts is None else parts[0] for parts in decoded)


def _evidence_parts(reference: str | None) -> tuple[str, str]:
    """(page, document reference) for one operand's evidence, as text.

    Returns the raw reference when it cannot be decoded, rather than an empty pair. The point of
    this column is that somebody can go and look; handing them nothing because the JSON changed
    shape defeats it.
    """
    if not reference:
        return ("", "")
    decoded = _decode_reference(reference)
    return decoded if decoded is not None else (UNREADABLE_REFERENCE, reference)


def _finding_row(finding: Finding) -> tuple[object, ...]:
    """One finding as a row, in `FINDING_COLUMNS` order.

    Keyed on **whether a calculation happened**, not on whether the outcome was a decision. The two
    are not the same: `Finding` requires a trace for a decision but *permits* one on an abstention,
    which is the ordinary shape of a `REVIEW_REQUIRED` raised partway through arithmetic — a unit it
    could not reconcile, say. When there is a trace, its comparison and tolerance are written,
    because the calculation genuinely ran and hiding it would leave a reviewer with an abstention and
    no idea how far it got.
    """
    trace = finding.trace
    abstained = is_abstention(finding.outcome)

    if trace is None:
        comparison: str = NOT_APPLICABLE if abstained else NONE_DECLARED
        tolerance: object = comparison
        unit: str = comparison
    else:
        comparison = trace.comparison
        tolerance = trace.tolerance if trace.tolerance is not None else NONE_DECLARED
        unit = trace.arithmetic_unit.value if trace.arithmetic_unit is not None else NONE_DECLARED

    return (
        finding.rule_id,
        finding.outcome.value,
        finding.severity.value,
        comparison,
        finding.delta if finding.delta is not None else (NOT_APPLICABLE if abstained else ""),
        tolerance,
        unit,
        finding.variant or "",
        finding.snapshot_id,
        finding.engine_version,
        _evidence_pages(finding),
        finding.reason,
        " | ".join(finding.notes),
    )


def _operand_rows(finding: Finding) -> list[tuple[object, ...]]:
    """One row per operand the calculation used, in `OPERAND_COLUMNS` order.

    A finding with no trace contributes no rows — nothing was read, and a row of blanks would
    suggest an operand was found and had no value. It still appears on the findings sheet, which is
    where a reader learns the check did not decide.

    An abstention that *does* carry a trace contributes its operands, for the same reason
    `_finding_row` writes its comparison: those operands were read. An abstention raised because two
    operands disagreed is one whose operands are the most useful thing in the file.
    """
    if finding.trace is None:
        return []
    rows: list[tuple[object, ...]] = []
    for operand in finding.trace.operands:
        page, document = _evidence_parts(operand.evidence_ref)
        rows.append((finding.rule_id, operand.name, operand.value, operand.source, page, document))
    return rows


def _write_sheet(
    workbook: Workbook, title: str, columns: Sequence[str], rows: Sequence[Sequence[object]]
) -> None:
    sheet = workbook.create_sheet(title)
    for index, name in enumerate(columns, start=1):
        heading = sheet.cell(row=1, column=index, value=name)
        heading.font = Font(bold=True)
        heading.number_format = TEXT_FORMAT
        # Wide enough to read without the reader resizing thirteen columns first. A guess, but a
        # cosmetic one — nothing here depends on it.
        sheet.column_dimensions[get_column_letter(index)].width = 22

    for offset, row in enumerate(rows, start=2):
        for index, value in enumerate(row, start=1):
            write_value(sheet.cell(row=offset, column=index), value)

    # Headings stay visible while scrolling, and the header row cannot be sorted into the data.
    sheet.freeze_panes = "A2"


def write_workbook(findings: Sequence[Finding]) -> bytes:
    """The findings as `.xlsx` bytes: every finding a row, every value text.

    Bytes rather than a path, so the caller decides where this goes — `storage/` hashes and stores
    artifacts, and a function that wrote to disk itself would either bypass that or duplicate it.

    The order given is the order written. Sorting here would make two exports of the same run
    differ from the redline, which is the document a reader has beside it.
    """
    if isinstance(findings, str) or not isinstance(findings, Sequence):
        raise TypeError("findings must be a sequence of Finding values")
    for finding in findings:
        if not isinstance(finding, Finding):
            raise TypeError("findings must contain only Finding values")

    workbook = Workbook()
    # A new Workbook comes with one sheet already; both sheets below are created explicitly, so the
    # default would otherwise sit at the front of the file as an empty "Sheet".
    workbook.remove(workbook.active)

    _write_sheet(workbook, FINDINGS_SHEET, FINDING_COLUMNS, [_finding_row(f) for f in findings])
    _write_sheet(
        workbook,
        OPERANDS_SHEET,
        OPERAND_COLUMNS,
        [row for finding in findings for row in _operand_rows(finding)],
    )

    out = BytesIO()
    workbook.save(out)
    return out.getvalue()
