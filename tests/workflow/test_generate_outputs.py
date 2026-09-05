"""The last stage: a reviewed package becomes a file somebody can be handed.

Verification for: `DatabaseStages.generate_outputs` in `workflow/stages.py` and
`reports/spreadsheet.py:write_stored_workbook` (#519).

Until this, checks ran, findings were recorded, and nothing turned them into a deliverable —
`write_workbook` had been finished and tested for months with no production caller, the same gap
#517 closed for crops and matching. With it, all six stages do work.

Two tests here are guards rather than demonstrations. `test_a_superseded_run_is_left_out_of_the_workbook`
protects a reviewer from a file containing two verdicts for one rule, and
`test_no_redline_is_produced_and_the_vocabulary_has_no_word_for_one` protects the boundary this
pipeline stops at: an annotated drawing needs each finding tied to a region of the sheet, which needs
semantic typing, which the pipeline deliberately does not do.
"""

from __future__ import annotations

import hashlib
import tempfile
from collections.abc import Iterator
from fractions import Fraction
from io import BytesIO
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import Engine, func, select
from sqlalchemy.orm import Session

from alembic import command
from app.db.session import session_factory
from app.models import OutputArtifact, OutputArtifactKind
from reports.spreadsheet import (
    FINDING_COLUMNS,
    FINDINGS_SHEET,
    NOT_RECORDED,
    OPERANDS_SHEET,
    TEXT_FORMAT,
)
from storage.local import LocalStore
from tests.app.postgres_fixture import alembic_config
from tests.workflow.test_stages import (
    _depth_operands,
    _live_findings,
    _project_depth_parameters,
    _publish_rulebook,
    _revision,
)
from workflow.stages import WORKBOOK_MEDIA_TYPE, DatabaseStages

pytest_plugins = ("tests.app.postgres_fixture",)

#: The first bytes of any `.xlsx`. A workbook is a zip archive, so this is what "a real file rather
#: than a row claiming one" looks like from outside openpyxl.
ZIP_SIGNATURE = b"PK\x03\x04"


def _upgrade(engine: Engine) -> None:
    config = alembic_config()
    config.attributes["database_url"] = engine.url.render_as_string(hide_password=False)
    command.upgrade(config, "head")


@pytest.fixture
def session(postgres_engine: Engine) -> Iterator[Session]:
    _upgrade(postgres_engine)
    opened = session_factory(postgres_engine)()
    try:
        yield opened
    finally:
        opened.close()


@pytest.fixture
def store() -> Iterator[LocalStore]:
    with tempfile.TemporaryDirectory() as directory:
        yield LocalStore(root=Path(directory), ticket_secret=b"a secret only this test knows")


def _checked(session: Session, store: LocalStore, *, depth: Fraction = Fraction(51, 2)):
    """A revision whose checks have run for real, with one rule reaching a decision.

    `CT-DEPTH-001` gets the project's parameters and a reviewer's reading, so the workbook under test
    contains at least one finding with a genuine calculation trace rather than only abstentions —
    the two take different paths through the writer.
    """
    revision = _revision(session)
    _publish_rulebook(session)
    _project_depth_parameters(session, revision)
    DatabaseStages(store, operands=_depth_operands(depth)).run_checks(session, revision.id)
    return revision


def _artifacts(session: Session, revision_id) -> list[OutputArtifact]:
    return list(
        session.execute(
            select(OutputArtifact).where(OutputArtifact.package_revision_id == revision_id)
        ).scalars()
    )


def _sheet(store: LocalStore, artifact: OutputArtifact, name: str):
    workbook = load_workbook(BytesIO(store.get(artifact.storage_key).read()))
    return workbook[name]


# ---------------------------------------------------------------------------
# The stage produces a real file
# ---------------------------------------------------------------------------


def test_a_reviewed_package_yields_a_persisted_findings_workbook(
    session: Session, store: LocalStore
) -> None:
    """The acceptance criterion: checks in, a deliverable out, recorded in the database.

    The row and the bytes are both asserted. A row pointing at nothing is the failure mode that
    matters — a reviewer follows the deliverable and finds it missing at the moment they need it.
    """
    revision = _checked(session, store)
    live = len(_live_findings(session, revision.id))

    result = DatabaseStages(store).generate_outputs(session, revision.id)

    assert result["implemented"] is True
    assert result["ran"] is True
    assert result["outputs"] == 1
    assert result["findings"] == live

    artifacts = _artifacts(session, revision.id)
    assert len(artifacts) == 1
    artifact = artifacts[0]
    assert artifact.kind == OutputArtifactKind.FINDINGS_WORKBOOK.value
    assert artifact.media_type == WORKBOOK_MEDIA_TYPE
    assert artifact.findings == live

    content = store.get(artifact.storage_key).read()
    assert content.startswith(ZIP_SIGNATURE), "the stored bytes are not a workbook"
    assert artifact.sha256 == hashlib.sha256(content).hexdigest()


def test_the_workbook_has_a_row_for_every_live_finding(session: Session, store: LocalStore) -> None:
    """Every finding, including the abstentions.

    A `NOT_FOUND` check has nothing to put in the value columns and the tempting thing is to leave
    the row out — which reproduces exactly the failure the whole system exists to prevent: a reader
    scanning for problems sees a short list and reads it as nothing wrong.
    """
    revision = _checked(session, store)
    live = _live_findings(session, revision.id)
    DatabaseStages(store).generate_outputs(session, revision.id)

    sheet = _sheet(store, _artifacts(session, revision.id)[0], FINDINGS_SHEET)

    assert sheet.max_row == len(live) + 1, "one heading row plus one row per finding"
    assert [cell.value for cell in sheet[1]] == list(FINDING_COLUMNS)
    outcomes = {sheet.cell(row=index, column=2).value for index in range(2, sheet.max_row + 1)}
    assert outcomes, "no outcomes were written"


def test_a_decision_carries_its_comparison_and_operands_into_the_file(
    session: Session, store: LocalStore
) -> None:
    """The point of the deliverable: a reviewer can see what was compared, not just the verdict.

    `CT-DEPTH-001` is given a reading that passes, so its row has a real comparison and its operands
    appear on the second sheet. Without this the export could contain nothing but abstentions and
    still look complete.
    """
    revision = _checked(session, store)
    DatabaseStages(store).generate_outputs(session, revision.id)
    artifact = _artifacts(session, revision.id)[0]

    findings = _sheet(store, artifact, FINDINGS_SHEET)
    rows = {
        findings.cell(row=index, column=1).value: [
            findings.cell(row=index, column=column).value
            for column in range(1, len(FINDING_COLUMNS) + 1)
        ]
        for index in range(2, findings.max_row + 1)
    }
    depth = rows["CT-DEPTH-001"]

    assert depth[1] == "PASS"
    comparison = depth[FINDING_COLUMNS.index("comparison")]
    assert comparison and "abstained" not in str(comparison)
    assert depth[FINDING_COLUMNS.index("rule_snapshot")], "the snapshot that decided is named"

    operands = _sheet(store, artifact, OPERANDS_SHEET)
    named = {operands.cell(row=index, column=1).value for index in range(2, operands.max_row + 1)}
    assert "CT-DEPTH-001" in named, "a decision contributed no operand rows"


def test_every_value_is_written_as_text(session: Session, store: LocalStore) -> None:
    """The module's one hard constraint, asserted on the file this stage actually writes.

    Give Excel `1/8` in a numeric cell and it stores `0.125`. The distinction between an eighth of an
    inch and a decimal approximation of one is what the units layer exists to preserve (ADR-0001),
    and losing it in the final step would make the export the least trustworthy artifact the system
    produces while looking the most familiar.
    """
    revision = _checked(session, store)
    DatabaseStages(store).generate_outputs(session, revision.id)
    sheet = _sheet(store, _artifacts(session, revision.id)[0], FINDINGS_SHEET)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            assert cell.number_format == TEXT_FORMAT, f"{cell.coordinate} is not a text cell"
            assert cell.value is None or isinstance(cell.value, str)


def test_the_columns_storage_never_kept_say_so_rather_than_being_blank(
    session: Session, store: LocalStore
) -> None:
    """`findings` does not store a decision's prose reason, its delta, its variant or its notes.

    Those live on the engine's value type and `record_finding` does not persist them. An empty cell
    would read as "there was nothing to say"; here there was something to say and nobody wrote it
    down, and a reader deciding whether to trust the column needs to be able to tell those apart.
    """
    revision = _checked(session, store)
    DatabaseStages(store).generate_outputs(session, revision.id)
    sheet = _sheet(store, _artifacts(session, revision.id)[0], FINDINGS_SHEET)

    difference = sheet.cell(row=2, column=FINDING_COLUMNS.index("difference") + 1).value
    variant = sheet.cell(row=2, column=FINDING_COLUMNS.index("variant") + 1).value

    assert difference == NOT_RECORDED
    assert variant == NOT_RECORDED


# ---------------------------------------------------------------------------
# What the workbook must not contain
# ---------------------------------------------------------------------------


def test_a_superseded_run_is_left_out_of_the_workbook(session: Session, store: LocalStore) -> None:
    """Re-running the checks replaces the results; the deliverable follows.

    A workbook carrying both runs would show a reviewer two verdicts for one rule with nothing
    saying which is in force — and both would look equally current, because a spreadsheet row has no
    way to whisper that it was superseded.
    """
    revision = _checked(session, store, depth=Fraction(51, 2))
    live = len(_live_findings(session, revision.id))
    # The same checks again with a reading that fails, which supersedes the first run.
    DatabaseStages(store, operands=_depth_operands(Fraction(101, 4))).run_checks(
        session, revision.id
    )

    result = DatabaseStages(store).generate_outputs(session, revision.id)

    assert (
        result["findings"] == live
    ), "the workbook counted superseded findings as well as live ones"
    total = session.execute(select(func.count()).select_from(OutputArtifact)).scalar_one()
    assert total == 1


def test_no_redline_is_produced_and_the_vocabulary_has_no_word_for_one(
    session: Session, store: LocalStore
) -> None:
    """**The hard stop.** Summary output only, and the schema cannot express anything else.

    An annotated drawing needs each finding tied to the region of the sheet it is about, which needs
    a candidate to have a meaning — and candidates are deliberately untyped until the real drawings
    (#274) and the vocabulary Q20 defers. A redline drawn from untyped candidates would put boxes on
    a drawing with nothing behind their placement, which is worse than no redline because it looks
    like evidence.

    Asserted on the enum as well as on the output, so adding a `redline` member is a deliberate act
    that fails this test and makes somebody read this docstring.
    """
    revision = _checked(session, store)
    DatabaseStages(store).generate_outputs(session, revision.id)

    kinds = {artifact.kind for artifact in _artifacts(session, revision.id)}

    assert kinds == {OutputArtifactKind.FINDINGS_WORKBOOK.value}
    assert [member.value for member in OutputArtifactKind] == ["findings_workbook"]


def test_a_revision_with_no_findings_produces_no_file(session: Session, store: LocalStore) -> None:
    """An empty workbook is a deliverable asserting a package was checked and found clean.

    Nothing ran, so nothing is produced, and the payload says which of the two it was.
    """
    revision = _revision(session)

    result = DatabaseStages(store).generate_outputs(session, revision.id)

    assert result["ran"] is True
    assert result["findings"] == 0
    assert result["outputs"] == 0
    assert "no live findings" in str(result["reason"])
    assert _artifacts(session, revision.id) == []


def test_regenerating_an_unchanged_revision_records_one_deliverable(
    session: Session, store: LocalStore
) -> None:
    """The key is content-addressed, and the table is append-only and unique on it.

    Running the stage twice on unchanged findings produces byte-identical bytes under the same key.
    Recording that twice would claim two deliverables where there is one, and the unique constraint
    would refuse the insert and take the transaction with it.
    """
    revision = _checked(session, store)
    first = DatabaseStages(store).generate_outputs(session, revision.id)

    second = DatabaseStages(store).generate_outputs(session, revision.id)

    assert first["outputs"] == 1
    assert second["outputs"] == 0
    assert second["already_recorded"] is True
    assert len(_artifacts(session, revision.id)) == 1


def test_without_a_store_it_says_so_rather_than_claiming_it_ran(session: Session) -> None:
    """No artifact store is a fact about this worker, not about the package."""
    result = DatabaseStages().generate_outputs(session, uuid4())

    assert result["implemented"] is True
    assert result["ran"] is False
    assert "no artifact store" in str(result["reason"])
