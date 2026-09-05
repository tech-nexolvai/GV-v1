"""The export must not quietly turn `1/8` into `0.125`.

Source: `AGENTS.md` §2.8; ADR-0001 · Design: `docs/DESIGN_PRODUCT.md` §3.3 ·
Verification: ``reports/spreadsheet.py``.

Two tests carry this file, and both read the workbook back rather than trusting what was written.

**Exact values survive as text.** A spreadsheet is the most eager type-coercer in the toolchain, and
a `1/8` stored as `0.125` looks entirely correct to a reader — which is why the assertion is on the
*stored type and number format*, not on the string handed to openpyxl.

**Abstentions are rows.** A missing row reads as nothing wrong, which is the failure the whole
abstention design exists to prevent. Asserted by counting rows against findings, so a filter added
later fails here rather than in front of a client.
"""

from __future__ import annotations

import json
from fractions import Fraction
from io import BytesIO
from typing import Final
from uuid import UUID

import pytest
from openpyxl import load_workbook

from reports.spreadsheet import (
    FINDING_COLUMNS,
    FINDINGS_SHEET,
    NOT_APPLICABLE,
    OPERAND_COLUMNS,
    OPERANDS_SHEET,
    TEXT_FORMAT,
    StoredFinding,
    exact_text,
    write_stored_workbook,
    write_workbook,
)
from units.measurement import Measurement, Unit
from verdict.finding import Finding
from verdict.outcomes import Outcome, Severity
from verdict.trace import CalculationTrace, TracedOperand

DOCUMENT: Final = UUID("11111111-1111-1111-1111-111111111111")


def _reference(page: int = 0) -> str:
    """The shape `evidence/gate.py` seals onto a verdict operand."""
    return json.dumps(
        {
            "document_version_id": str(DOCUMENT),
            "page": page,
            "polygon": [["0.10", "0.20"], ["0.50", "0.20"], ["0.50", "0.40"], ["0.10", "0.40"]],
            "space": "stored",
        },
        sort_keys=True,
        separators=(",", ":"),
    )


def _inches(numerator: int, denominator: int = 1, raw: str | None = None) -> Measurement:
    return Measurement(Fraction(numerator, denominator), Unit.INCH, raw)


def _trace(
    *,
    operands: tuple[TracedOperand, ...] | None = None,
    tolerance: Measurement | None = None,
    unit: Unit | None = Unit.INCH,
) -> CalculationTrace:
    return CalculationTrace(
        operation="sum_within_tolerance",
        operands=(
            operands
            if operands is not None
            else (
                TracedOperand(
                    name="countertop_width",
                    value=_inches(97, 8, raw='12 1/8"'),
                    source="SHOP",
                    evidence_ref=_reference(),
                ),
            )
        ),
        intermediates=(("expected_width", "12 1/8"),),
        comparison="12 1/8 in vs 12 in",
        tolerance=tolerance,
        arithmetic_unit=unit,
        outcome=Outcome.FAIL,
        engine_version="test",
        operation_version="1",
    )


def _finding(
    rule_id: str = "CT-WIDTH-001",
    outcome: Outcome = Outcome.FAIL,
    *,
    trace: CalculationTrace | None = None,
    delta: Measurement | None = None,
    refs: tuple[str, ...] = (),
    variant: str | None = None,
    notes: tuple[str, ...] = (),
) -> Finding:
    decided = outcome in (Outcome.PASS, Outcome.FAIL)
    return Finding(
        rule_id=rule_id,
        outcome=outcome,
        severity=Severity.CRITICAL,
        reason="the countertop is an eighth of an inch wider than the cabinets beneath it",
        snapshot_id="sha256:0123456789abcdef",
        engine_version="verdict-1.2.3",
        trace=trace if trace is not None else (_trace() if decided else None),
        delta=delta,
        evidence_refs=refs,
        variant=variant,
        notes=notes,
    )


def _sheet(data: bytes, title: str) -> list[list[object]]:
    """Every row of one sheet, as read back from the saved file."""
    workbook = load_workbook(BytesIO(data))
    return [list(row) for row in workbook[title].iter_rows(values_only=True)]


def _cells(data: bytes, title: str) -> list[list[object]]:
    """The cell objects, so stored type and number format can be asserted."""
    workbook = load_workbook(BytesIO(data))
    return [list(row) for row in workbook[title].iter_rows()]


# ---------------------------------------------------------------------------
# Exact values survive
# ---------------------------------------------------------------------------


def test_an_eighth_of_an_inch_is_not_stored_as_a_decimal() -> None:
    """The one thing this module exists for.

    `0.125` in a cell looks entirely correct, which is why this asserts on the stored value and not
    on whether the number happens to be equal.
    """
    data = write_workbook([_finding()])
    rows = _sheet(data, OPERANDS_SHEET)

    values = [row[OPERAND_COLUMNS.index("value")] for row in rows[1:]]
    assert any("97/8 in" in str(value) for value in values)
    assert not any("12.125" in str(value) or "0.125" in str(value) for value in values)


def test_every_value_cell_is_stored_as_text_and_formatted_as_text() -> None:
    """Both halves, because either one alone leaves the door open.

    A numeric cell converts on write. A string cell left on `General` converts the first time a
    consumer's tooling re-saves the file, which is worse — the export was correct when it left here
    and is wrong by the time anyone disputes it.
    """
    data = write_workbook([_finding()])

    for title in (FINDINGS_SHEET, OPERANDS_SHEET):
        for row in _cells(data, title):
            for cell in row:
                if cell.value is None:
                    continue
                assert isinstance(cell.value, str), f"{title}!{cell.coordinate} is not text"
                assert cell.number_format == TEXT_FORMAT, f"{title}!{cell.coordinate} may convert"


def test_a_measurement_keeps_its_unit_and_what_the_drawing_said() -> None:
    """A number without its unit is the other half of the same mistake, and the source token is how
    a reader checks the reading rather than only the verdict."""
    rendered = exact_text(_inches(97, 8, raw='12 1/8"'))

    assert "97/8 in" in rendered
    assert '12 1/8"' in rendered


def test_a_fraction_renders_as_a_fraction() -> None:
    assert exact_text(Fraction(1, 8)) == "1/8"


def test_a_whole_number_does_not_acquire_a_denominator() -> None:
    assert exact_text(Fraction(6010)) == "6010"


# ---------------------------------------------------------------------------
# Abstentions are rows
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "outcome",
    [Outcome.NOT_FOUND, Outcome.REVIEW_REQUIRED, Outcome.NO_APPLICABLE_RULE],
    ids=lambda outcome: outcome.value,
)
def test_an_abstention_appears_as_a_row(outcome: Outcome) -> None:
    """A missing row reads as nothing wrong. That is the failure `NO_APPLICABLE_RULE` was invented
    to stop, and an export that dropped these would reintroduce it in the format the client
    actually reads."""
    data = write_workbook([_finding("CT-DEPTH-001", outcome)])
    rows = _sheet(data, FINDINGS_SHEET)

    assert len(rows) == 2, "one heading row and one finding row"
    assert rows[1][FINDING_COLUMNS.index("outcome")] == outcome.value
    assert rows[1][FINDING_COLUMNS.index("check")] == "CT-DEPTH-001"


def test_an_abstentions_value_columns_say_why_they_are_empty() -> None:
    """A blank cell is ambiguous between "nothing to report", "the export dropped it" and "the
    check was never run", and those want different responses from a reader."""
    data = write_workbook([_finding(outcome=Outcome.NOT_FOUND)])
    row = _sheet(data, FINDINGS_SHEET)[1]

    assert "abstained" in str(row[FINDING_COLUMNS.index("comparison")])
    assert "abstained" in str(row[FINDING_COLUMNS.index("tolerance")])


def test_every_finding_reaches_the_workbook() -> None:
    """Counted against the input, so a filter added later fails here rather than in front of a
    client."""
    findings = [
        _finding("CT-A"),
        _finding("CT-B", Outcome.PASS),
        _finding("CT-C", Outcome.NOT_FOUND),
        _finding("CT-D", Outcome.REVIEW_REQUIRED),
        _finding("CT-E", Outcome.NO_APPLICABLE_RULE),
    ]

    rows = _sheet(write_workbook(findings), FINDINGS_SHEET)

    assert len(rows) == len(findings) + 1
    assert [row[0] for row in rows[1:]] == ["CT-A", "CT-B", "CT-C", "CT-D", "CT-E"]


def test_an_abstention_contributes_no_operand_rows() -> None:
    """Nothing was read. A row of blanks would suggest an operand was found and had no value — and
    the finding is on the findings sheet either way, which is where "did not decide" belongs."""
    rows = _sheet(write_workbook([_finding(outcome=Outcome.NOT_FOUND)]), OPERANDS_SHEET)

    assert len(rows) == 1, "the heading row only"


# ---------------------------------------------------------------------------
# The layout is stable
# ---------------------------------------------------------------------------


def test_the_column_headings_are_exactly_the_declared_layout() -> None:
    """Downstream consumers index by position. A column inserted in the middle shifts every one
    after it, and a reader gets tolerances under the severity heading with no reason to doubt
    them."""
    data = write_workbook([_finding()])

    assert tuple(_sheet(data, FINDINGS_SHEET)[0]) == FINDING_COLUMNS
    assert tuple(_sheet(data, OPERANDS_SHEET)[0]) == OPERAND_COLUMNS


def test_the_same_findings_produce_the_same_rows_twice() -> None:
    """The report is reconciled against a previous copy, and a row order that moved between runs
    would read as a change that did not happen.

    The bytes themselves are not compared: `.xlsx` is a zip, and its entries carry timestamps.
    """
    findings = [_finding("CT-A"), _finding("CT-B", Outcome.NOT_FOUND), _finding("CT-C")]

    first = write_workbook(findings)
    second = write_workbook(findings)

    for title in (FINDINGS_SHEET, OPERANDS_SHEET):
        assert _sheet(first, title) == _sheet(second, title)


def test_the_input_order_is_preserved() -> None:
    """Not sorted here. Sorting would make the workbook disagree with the redline, which is the
    document a reader has open beside it."""
    findings = [_finding("CT-Z"), _finding("CT-A"), _finding("CT-M")]

    rows = _sheet(write_workbook(findings), FINDINGS_SHEET)

    assert [row[0] for row in rows[1:]] == ["CT-Z", "CT-A", "CT-M"]


# ---------------------------------------------------------------------------
# Provenance a reader can follow
# ---------------------------------------------------------------------------


def test_an_operand_row_points_at_the_page_its_value_came_from() -> None:
    """One-based, because the reader is holding a drawing set rather than an array."""
    operands = (
        TracedOperand(
            name="countertop_width",
            value=_inches(97, 8),
            source="SHOP",
            evidence_ref=_reference(page=3),
        ),
    )
    data = write_workbook([_finding(trace=_trace(operands=operands))])
    row = _sheet(data, OPERANDS_SHEET)[1]

    assert row[OPERAND_COLUMNS.index("evidence_page")] == "4"
    assert row[OPERAND_COLUMNS.index("evidence_uri")] == str(DOCUMENT)
    assert row[OPERAND_COLUMNS.index("source")] == "SHOP"


def test_an_undecodable_reference_is_reported_rather_than_dropped() -> None:
    """A reference the export could not read is a fact about the export. Silently blanking it leaves
    the row looking like a finding with no evidence, which is a different and more comfortable
    problem than the one that actually occurred."""
    operands = (
        TracedOperand(
            name="countertop_width", value=_inches(97, 8), source="SHOP", evidence_ref="not json"
        ),
    )
    data = write_workbook([_finding(trace=_trace(operands=operands), refs=("not json",))])

    operand_row = _sheet(data, OPERANDS_SHEET)[1]
    assert operand_row[OPERAND_COLUMNS.index("evidence_page")] == "unreadable reference"
    assert operand_row[OPERAND_COLUMNS.index("evidence_uri")] == "not json"

    finding_row = _sheet(data, FINDINGS_SHEET)[1]
    assert "unreadable reference" in str(finding_row[FINDING_COLUMNS.index("evidence_pages")])


def test_the_row_records_the_snapshot_and_engine_that_produced_it() -> None:
    """Without them a row cannot be reproduced, which is the whole reason snapshots exist
    (ADR-0005) — and a spreadsheet outlives the run that made it."""
    row = _sheet(write_workbook([_finding()]), FINDINGS_SHEET)[1]

    assert row[FINDING_COLUMNS.index("rule_snapshot")] == "sha256:0123456789abcdef"
    assert row[FINDING_COLUMNS.index("engine_version")] == "verdict-1.2.3"


def test_a_declared_tolerance_is_written_exactly() -> None:
    tolerance = _inches(1, 16)
    row = _sheet(write_workbook([_finding(trace=_trace(tolerance=tolerance))]), FINDINGS_SHEET)[1]

    assert "1/16 in" in str(row[FINDING_COLUMNS.index("tolerance")])


def test_a_rule_with_no_tolerance_says_so_rather_than_leaving_a_blank() -> None:
    """An exact-match rule (Q2) declares no tolerance. A blank cell there is indistinguishable from
    a tolerance the export failed to write."""
    row = _sheet(write_workbook([_finding(trace=_trace(tolerance=None))]), FINDINGS_SHEET)[1]

    assert row[FINDING_COLUMNS.index("tolerance")] == "none declared"


def test_the_variant_and_notes_reach_the_row() -> None:
    """The variant is which layout's tolerance applied, and a note is something a reviewer should
    see that did not change the outcome — both invisible in the outcome column alone."""
    row = _sheet(
        write_workbook(
            [_finding(variant="back_left_right", notes=("company standard overridden",))]
        ),
        FINDINGS_SHEET,
    )[1]

    assert row[FINDING_COLUMNS.index("variant")] == "back_left_right"
    assert "company standard overridden" in str(row[FINDING_COLUMNS.index("notes")])


# ---------------------------------------------------------------------------
# Refusals
# ---------------------------------------------------------------------------


def test_a_workbook_of_no_findings_still_has_its_headings() -> None:
    """An empty run is a legitimate result, and a file with no header row is one a consumer cannot
    read at all."""
    data = write_workbook([])

    assert tuple(_sheet(data, FINDINGS_SHEET)[0]) == FINDING_COLUMNS
    assert len(_sheet(data, FINDINGS_SHEET)) == 1


def test_something_that_is_not_a_finding_is_refused() -> None:
    with pytest.raises(TypeError, match="only Finding values"):
        write_workbook(["CT-WIDTH-001"])  # type: ignore[list-item]


def test_a_bare_finding_is_refused_rather_than_iterated() -> None:
    """A `Finding` is not a sequence, and a string is — the check exists so neither is silently
    treated as a list of findings."""
    with pytest.raises(TypeError, match="sequence of Finding"):
        write_workbook("CT-WIDTH-001")  # type: ignore[arg-type]


def test_a_reference_without_its_document_is_not_reported_as_a_page() -> None:
    """ "Page 1" of which drawing?

    A reference carrying `page` and no `document_version_id` decodes happily. Reporting it as
    provenance puts an unfollowable citation in the column a reader uses to go and check — they
    look, fail to find it, and doubt the finding rather than the export.
    """
    partial = json.dumps({"page": 0}, separators=(",", ":"))
    operands = (
        TracedOperand(
            name="countertop_width", value=_inches(97, 8), source="SHOP", evidence_ref=partial
        ),
    )
    data = write_workbook([_finding(trace=_trace(operands=operands), refs=(partial,))])

    finding_row = _sheet(data, FINDINGS_SHEET)[1]
    assert finding_row[FINDING_COLUMNS.index("evidence_pages")] == "unreadable reference"

    operand_row = _sheet(data, OPERANDS_SHEET)[1]
    assert operand_row[OPERAND_COLUMNS.index("evidence_page")] == "unreadable reference"
    assert operand_row[OPERAND_COLUMNS.index("evidence_uri")] == partial


def test_both_sheets_agree_about_what_is_readable() -> None:
    """One decoder for both callers.

    Two copies of "what counts as a readable reference" is how the operand sheet and the findings
    sheet come to disagree about the same reference, and a reader comparing them has no way to tell
    which is right.
    """
    for reference in (
        '{"page": 0}',
        '{"document_version_id": "x"}',
        "not json",
        '{"page": "1"}',
        # Shape-valid and still not a citation: no document can be retrieved by an empty id, and a
        # negative page renders as page 0 of a set whose first sheet is 1.
        '{"page": 0, "document_version_id": ""}',
        '{"page": 0, "document_version_id": "   "}',
        '{"page": -1, "document_version_id": "11111111-1111-1111-1111-111111111111"}',
        '{"page": true, "document_version_id": "11111111-1111-1111-1111-111111111111"}',
    ):
        operands = (
            TracedOperand(name="w", value=_inches(1), source="SHOP", evidence_ref=reference),
        )
        data = write_workbook([_finding(trace=_trace(operands=operands), refs=(reference,))])

        from_findings = _sheet(data, FINDINGS_SHEET)[1][FINDING_COLUMNS.index("evidence_pages")]
        from_operands = _sheet(data, OPERANDS_SHEET)[1][OPERAND_COLUMNS.index("evidence_page")]

        assert from_findings == from_operands == "unreadable reference", reference


def test_an_abstention_that_carries_a_trace_shows_what_it_computed() -> None:
    """`Finding` requires a trace for a decision but permits one on an abstention, which is the
    ordinary shape of a REVIEW_REQUIRED raised partway through arithmetic.

    The columns follow the calculation, not the outcome. Hiding a trace that exists would leave a
    reviewer with an abstention and no idea how far the check got — and an abstention raised because
    two readings disagreed is one whose readings are the most useful thing in the file.
    """
    data = write_workbook([_finding(outcome=Outcome.REVIEW_REQUIRED, trace=_trace())])

    finding_row = _sheet(data, FINDINGS_SHEET)[1]
    assert finding_row[FINDING_COLUMNS.index("outcome")] == Outcome.REVIEW_REQUIRED.value
    assert finding_row[FINDING_COLUMNS.index("comparison")] == "12 1/8 in vs 12 in"
    assert "abstained" not in str(finding_row[FINDING_COLUMNS.index("comparison")])

    assert len(_sheet(data, OPERANDS_SHEET)) == 2, "its operands were read, so they are listed"


# ---------------------------------------------------------------------------
# The workbook a worker writes, from stored rows
# ---------------------------------------------------------------------------


def _stored(**overrides: object) -> StoredFinding:
    """A stored finding with a real calculation trace, in the shape `app/verdicts/trace.py` writes."""
    defaults: dict[str, object] = {
        "rule_id": "CT-DEPTH-001",
        "outcome": "PASS",
        "severity": "CRITICAL",
        "snapshot_id": "a" * 64,
        "engine_version": "verdict/1",
        "trace": {
            "operation": "sum_within_tolerance",
            "operands": [
                {
                    "name": "countertop_depth",
                    "value": "25 1/2 in",
                    "source": "USER_INPUT",
                    "evidence_ref": None,
                }
            ],
            "intermediates": [],
            "comparison": "25 1/2 == 25 1/2",
            "tolerance": None,
            "arithmetic_unit": "in",
            "outcome": "PASS",
            "engine_version": "verdict/1",
            "operation_version": 1,
        },
    }
    defaults.update(overrides)
    return StoredFinding(**defaults)  # type: ignore[arg-type]


def test_a_stored_fraction_survives_as_the_text_it_was_written_as() -> None:
    """`25 1/2` reaches the cell as `25 1/2`, not as `25.5`.

    The stored value is already exact text — `render_value` produced it — so the only way to lose it
    here is to let openpyxl decide the cell is numeric. This is the same guarantee
    `test_every_measurement_is_a_text_cell` makes for the engine-value path, asserted separately
    because the two writers reach the cell by different routes.
    """
    sheet = load_workbook(BytesIO(write_stored_workbook([_stored()])))[OPERANDS_SHEET]

    value = sheet.cell(row=2, column=OPERAND_COLUMNS.index("value") + 1)

    assert value.value == "25 1/2 in"
    assert value.number_format == TEXT_FORMAT


def test_an_abstention_carries_its_stored_reason() -> None:
    """The reason an abstention records is the one thing storage does keep, and it must appear.

    "No dimension was read for `cutout_width`" sends somebody to the drawing. Dropping it in the
    export sends them to us.
    """
    finding = _stored(
        outcome="NOT_FOUND",
        trace={
            "cause": "missing_operand",
            "reason": "no dimension was read for cutout_width",
            "outcome": "NOT_FOUND",
        },
    )

    sheet = load_workbook(BytesIO(write_stored_workbook([finding])))[FINDINGS_SHEET]

    assert sheet.cell(row=2, column=FINDING_COLUMNS.index("reason") + 1).value == (
        "no dimension was read for cutout_width"
    )
    assert sheet.cell(row=2, column=FINDING_COLUMNS.index("comparison") + 1).value == NOT_APPLICABLE


def test_a_stored_abstention_contributes_no_operand_rows() -> None:
    """Nothing was read, and a row of blanks would suggest an operand was found and had no value."""
    finding = _stored(
        outcome="NOT_FOUND",
        trace={"cause": "missing_operand", "reason": "nothing was read", "outcome": "NOT_FOUND"},
    )

    sheet = load_workbook(BytesIO(write_stored_workbook([finding])))[OPERANDS_SHEET]

    assert sheet.max_row == 1, "only the heading row"


def test_a_trace_of_an_unexpected_shape_is_rendered_rather_than_raising() -> None:
    """`trace` is `JSONB`, so an older writer's row can carry fields this reader did not expect.

    The export's job is to show a reviewer what was recorded. Raising would make one unfamiliar row
    withhold the whole deliverable, including the rows that are fine.
    """
    finding = _stored(trace={"operands": "not a list", "comparison": 7})

    sheet = load_workbook(BytesIO(write_stored_workbook([finding])))[FINDINGS_SHEET]

    assert sheet.cell(row=2, column=FINDING_COLUMNS.index("comparison") + 1).value == "7"


def test_the_two_writers_agree_on_their_columns() -> None:
    """Both sheets are written through `_write_sheet` with the same column tuples.

    Asserted so the stored-row export and the engine-value export stay comparable: a reader diffing
    one against the other is entitled to assume column five means the same thing in both.
    """
    stored = load_workbook(BytesIO(write_stored_workbook([_stored()])))
    assert [cell.value for cell in stored[FINDINGS_SHEET][1]] == list(FINDING_COLUMNS)
    assert [cell.value for cell in stored[OPERANDS_SHEET][1]] == list(OPERAND_COLUMNS)
